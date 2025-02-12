"""Data Dictionary."""
# Standard Python Libraries
import argparse
import ast
import os
import sys

# Third-Party Libraries
import astor
from openpyxl import Workbook, load_workbook
import pandas as pd
import psycopg2

################################################################################
# 1) SQL "template" that takes a {schema_name} placeholder for fromSql
################################################################################
POSTGRES_DATA_DICTIONARY_SQL_TEMPLATE = r"""
------------------------------------------------------------------------------------
-- Data Dictionary Dump:
-- This SQL script will dump table, column, key, and description design related
-- metadata so that you can copy-paste or export to Excel as a Data Dictionary.
------------------------------------------------------------------------------------
-- Platform:          PostgreSQL
-- Author:            DataResearchLabs
-- GitHub:            https://github.com/DataResearchLabs/sql_scripts
-- YouTube Tutorials: https://www.youtube.com/channel/UCQciXv3xaBykeUFc04GxSXA
----------------------------------------------------------------------------------
WITH vars AS (
    SELECT
        '{schema_name}'::text AS v_SchemaName,
        'NO'                  AS v_TablesOnly
)
, baseTbl AS (
    SELECT
          table_schema AS SchemaName
        , table_catalog
        , table_type
        , table_name
    FROM INFORMATION_SCHEMA.TABLES
    WHERE table_schema = (SELECT v_SchemaName FROM vars)
      AND (
            (table_type = 'BASE TABLE')
         OR ((SELECT v_TablesOnly FROM vars) = 'NO')
          )
)
, metadata AS (
    SELECT
          bt.SchemaName                AS schema_nm
        , bt.table_name                AS table_nm
        , CASE
            WHEN bt.table_type = 'BASE TABLE' THEN 'TBL'
            WHEN bt.table_type = 'VIEW'       THEN 'VW'
            ELSE 'UK'
          END                           AS obj_typ
        , tut.ordinal_position         AS ord
        , tut.column_name              AS column_nm
        , CONCAT(
            COALESCE(tut.data_type, 'unknown'),
            CASE
                WHEN tut.data_type IN ('varchar','char') THEN
                    CONCAT('(', CAST(tut.CHARACTER_MAXIMUM_LENGTH AS varchar(10)), ')')
                WHEN tut.data_type IN('date','time') THEN '(3)'
                WHEN tut.data_type = 'datetime' THEN '(8)'
                WHEN tut.data_type = 'timestamp' THEN '(4)'
                WHEN tut.data_type IN('bigint','integer','smallint') THEN
                    CONCAT('(', CAST(tut.NUMERIC_PRECISION AS varchar(10)), ')')
                WHEN tut.data_type = 'decimal' THEN
                    CONCAT('(', CAST(tut.NUMERIC_PRECISION AS varchar(10)), ',',
                               CAST(tut.NUMERIC_SCALE AS varchar(10)), ')')
                WHEN tut.CHARACTER_MAXIMUM_LENGTH IS NOT NULL THEN
                    CONCAT('(', CAST(tut.CHARACTER_MAXIMUM_LENGTH AS varchar(10)), ')')
                WHEN tut.DATETIME_PRECISION IS NOT NULL THEN
                    CONCAT('(', CAST(tut.DATETIME_PRECISION AS varchar(10)), ')')
                WHEN tut.NUMERIC_PRECISION IS NOT NULL
                     AND tut.NUMERIC_SCALE IS NULL THEN
                    CONCAT('(', CAST(tut.NUMERIC_PRECISION AS varchar(10)), ')')
                WHEN tut.NUMERIC_PRECISION IS NOT NULL
                     AND tut.NUMERIC_SCALE IS NOT NULL THEN
                    CONCAT('(', CAST(tut.NUMERIC_PRECISION AS varchar(10)), ',',
                               CAST(tut.NUMERIC_SCALE AS varchar(10)), ')')
                ELSE
                    ''
            END
          )                              AS data_typ
        , CASE
            WHEN tut.is_nullable = 'YES' THEN 'NULL'
            ELSE 'NOT NULL'
          END                            AS nullable
    FROM INFORMATION_SCHEMA.COLUMNS tut
    INNER JOIN baseTbl bt
        ON bt.table_catalog = tut.table_catalog
       AND bt.table_name    = tut.table_name
)
, meta_for_keys AS (
    SELECT
          schema_nm
        , table_nm
        , column_nm
        , STRING_AGG(is_key, ',' ORDER BY is_key) AS is_key
    FROM (
        SELECT
              cons.table_schema    AS schema_nm
            , cons.table_name      AS table_nm
            , kcu.column_name      AS column_nm
            , CASE
                WHEN cons.constraint_type = 'PRIMARY KEY' THEN 'PK'
                WHEN cons.constraint_type = 'UNIQUE'      THEN 'UK'
                WHEN cons.constraint_type = 'FOREIGN KEY' THEN 'FK'
                ELSE 'X'
              END AS is_key
        FROM INFORMATION_SCHEMA.table_constraints  cons
        INNER JOIN INFORMATION_SCHEMA.key_column_usage kcu
            ON cons.table_schema    = kcu.table_schema
           AND cons.table_name      = kcu.table_name
           AND cons.constraint_name = kcu.constraint_name
        WHERE cons.table_schema = (SELECT v_SchemaName FROM vars)
          AND cons.table_name   IN (SELECT DISTINCT table_name FROM baseTbl)
          AND cons.constraint_type IN ('PRIMARY KEY','FOREIGN KEY','UNIQUE')
        GROUP BY
              cons.table_schema
            , cons.table_name
            , kcu.column_name
            , cons.constraint_type
    ) t
    GROUP BY schema_nm, table_nm, column_nm
)
, col_comm AS (
    SELECT
          c.table_schema AS schema_nm
        , c.table_name   AS table_nm
        , c.column_name  AS column_nm
        , pgd.description AS column_descr
    FROM pg_catalog.pg_statio_all_tables st
    INNER JOIN pg_catalog.pg_description pgd
        ON pgd.objoid = st.relid
    INNER JOIN information_schema.columns c
        ON pgd.objsubid   = c.ordinal_position
       AND c.table_schema = st.schemaname
       AND c.table_name   = st.relname
    WHERE c.table_schema IN (SELECT v_SchemaName FROM vars)
      AND c.table_name   IN (SELECT DISTINCT table_name FROM baseTbl)
)
SELECT
      md.schema_nm
    , md.table_nm
    , md.obj_typ
    , md.ord
    , COALESCE(pk.is_key, ' ') AS is_key
    , md.column_nm
    , md.data_typ
    , md.nullable
    , c.column_descr
FROM metadata md
LEFT JOIN meta_for_keys pk
       ON pk.schema_nm  = md.schema_nm
      AND pk.table_nm   = md.table_nm
      AND pk.column_nm  = md.column_nm
LEFT JOIN col_comm c
       ON c.schema_nm   = md.schema_nm
      AND c.table_nm    = md.table_nm
      AND c.column_nm   = md.column_nm
ORDER BY
      md.schema_nm
    , md.table_nm
    , md.ord
"""

################################################################################
# 2) Main script logic, with four modes: fromExcel, toExcel, fromSql, toSql
################################################################################


def main(
    models_path,
    excel_path,
    output_path,
    from_excel,
    to_excel,
    from_sql,
    to_sql,
    db_host,
    db_port,
    db_name,
    db_user,
    db_password,
    db_schema,
):
    """Run main function."""
    if from_sql:
        # ============ (A) Export from Postgres -> CSV using embedded query
        sql_query = POSTGRES_DATA_DICTIONARY_SQL_TEMPLATE.format(schema_name=db_schema)
        export_data_dictionary_to_csv(
            query=sql_query,
            csv_path=excel_path,
            host=db_host,
            port=db_port,
            dbname=db_name,
            user=db_user,
            password=db_password,
        )

    elif to_sql:
        # ============ (B) Update Postgres column descriptions from the Excel
        update_postgres_column_comments_from_excel(
            excel_path=excel_path,
            host=db_host,
            port=db_port,
            dbname=db_name,
            user=db_user,
            password=db_password,
            schema=db_schema,
        )

    elif from_excel:
        print("Updating Models file " + models_path + " from " + excel_path)
        # ============ (C) Update Django models.py from Excel
        # We load the Excel/CSV with columns:
        # schema_nm, table_nm, obj_typ, ord, is_key, column_nm, data_typ, nullable, column_descr
        help_dict = load_help_dict_from_excel(excel_path)
        new_source = add_help_text_from_excel(models_path, help_dict)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(new_source)
        print(f"[INFO] Updated {output_path} from Excel/CSV: {excel_path}")

    elif to_excel:
        print("Updating Excel file " + excel_path + " from " + models_path)
        # ============ (D) Update/Create the Excel/CSV from Django models.py
        help_dict = extract_help_text_from_source(models_path)
        update_excel_from_help_dict(excel_path, help_dict)


################################################################################
# 3) fromSql: Export to CSV
################################################################################


def export_data_dictionary_to_csv(query, csv_path, host, port, dbname, user, password):
    """Export data dictionary."""
    conn = psycopg2.connect(
        host=host, port=port, dbname=dbname, user=user, password=password
    )
    df = pd.read_sql(query, conn)
    df.to_csv(csv_path, index=False)
    conn.close()
    print(f"[INFO] Exported data dictionary to CSV: {csv_path}")


################################################################################
# 4) toSql: Update Postgres from Excel
################################################################################


def update_postgres_column_comments_from_excel(
    excel_path, host, port, dbname, user, password, schema="public"
):
    """
    We expect columns in the file.

    schema_nm, table_nm, obj_typ, ord, is_key, column_nm, data_typ, nullable, column_descr
    'column_descr' is used as the help text for COMMENT ON COLUMN.
    """
    # Load either Excel or CSV; we'll do a quick sniff:
    if excel_path.lower().endswith(".csv"):
        df = pd.read_csv(excel_path)
    else:
        df = pd.read_excel(excel_path, engine="openpyxl")

    conn = psycopg2.connect(
        host=host, port=port, dbname=dbname, user=user, password=password
    )
    cur = conn.cursor()

    for _, row in df.iterrows():
        # We only *really* need table_nm, column_nm, column_descr
        table = str(row.get("table_nm", "")).strip()
        column = str(row.get("column_nm", "")).strip()
        comment = str(row.get("column_descr", "")).strip()

        if not table or not column:
            continue

        if comment and comment.lower() != "none":
            sql = f'COMMENT ON COLUMN "{schema}"."{table}"."{column}" IS %s'
            cur.execute(sql, (comment,))
        else:
            sql = f'COMMENT ON COLUMN "{schema}"."{table}"."{column}" IS NULL'
            cur.execute(sql)

    conn.commit()
    cur.close()
    conn.close()
    print(f"[INFO] Updated Postgres from {excel_path} using column_descr.")


################################################################################
# 5) fromExcel: Update Django models from Excel
################################################################################


REQUIRED_COLUMNS = [
    "schema_nm",
    "table_nm",
    "obj_typ",
    "ord",
    "is_key",
    "column_nm",
    "data_typ",
    "nullable",
    "column_descr",
]


def load_help_dict_from_excel(xlsx_path):
    """
    Load help dictionary.

    1) Ensure the XLSX file exists. If not, create it with REQUIRED_COLUMNS in row 1.
    2) Read the XLSX (or CSV) into a DataFrame, making sure we have all REQUIRED_COLUMNS.
    3) Build and return a dict of shape:
         help_dict = {
            <table_nm>: {
                <column_nm>: <column_descr>,
                ...
            },
            ...
         }
    """
    # If the file doesn't exist, create an empty Excel with headers.
    if not os.path.exists(xlsx_path):
        # Create a new workbook with the required columns as the first row.
        wb = Workbook()
        ws = wb.active
        ws.append(REQUIRED_COLUMNS)
        wb.save(xlsx_path)
        print(f"[INFO] Created new XLSX file with headers: {xlsx_path}")

    # Now read the Excel or CSV.
    if xlsx_path.lower().endswith(".csv"):
        df = pd.read_csv(xlsx_path)
    else:
        df = pd.read_excel(xlsx_path, engine="openpyxl")

    # Ensure all required columns exist (fill missing with empty strings).
    for col in REQUIRED_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    # Build a help_dict mapping {table_nm: {column_nm: column_descr}}
    help_dict = {}
    for _, row in df.iterrows():
        tbl = str(row["table_nm"]).strip()
        col = str(row["column_nm"]).strip()
        descr = str(row["column_descr"]).strip()

        # Skip rows with no table or column name
        if not tbl or not col:
            continue

        help_dict.setdefault(tbl, {})[col] = descr

    return help_dict


def add_help_text_from_excel(models_path, help_dict):
    """
    Parse the models.py file.

    Then for each class_name in help_dict, add help_text to the relevant fields.
    """
    with open(models_path, encoding="utf-8") as f:
        source = f.read()

    tree = ast.parse(source)
    for node in tree.body:
        if isinstance(node, ast.ClassDef):
            class_name = node.name
            # If the class_name is in help_dict:
            if is_django_model_class(node) and class_name in help_dict:
                class_help = help_dict[class_name]
                for body_node in node.body:
                    if (
                        isinstance(body_node, ast.Assign)
                        and len(body_node.targets) == 1
                        and isinstance(body_node.targets[0], ast.Name)
                    ):
                        field_name = body_node.targets[0].id
                        if field_name in class_help:
                            add_help_text_if_missing(
                                body_node.value, class_help[field_name]
                            )

    return astor.to_source(tree)


################################################################################
# 6) toExcel: Export Django models to Excel/CSV using the same columns
################################################################################


def extract_help_text_from_source(models_path):
    """
    Return a dict in shape.

      help_dict = {
         modelName: { fieldName: helpText, ... },
         ...
      }
    """
    with open(models_path, encoding="utf-8") as f:
        source = f.read()

    tree = ast.parse(source)
    help_dict = {}
    for node in tree.body:
        if isinstance(node, ast.ClassDef):
            class_name = node.name
            if is_django_model_class(node):
                class_help = {}
                for body_node in node.body:
                    if (
                        isinstance(body_node, ast.Assign)
                        and len(body_node.targets) == 1
                        and isinstance(body_node.targets[0], ast.Name)
                    ):
                        field_name = body_node.targets[0].id
                        help_text = get_help_text_from_field(body_node.value)
                        if help_text:
                            class_help[field_name] = help_text
                if class_help:
                    help_dict[class_name] = class_help
    return help_dict


def update_excel_from_help_dict(excel_path, help_dict):
    """
    We want to produce columns.

    schema_nm, table_nm, obj_typ, ord, is_key, column_nm, data_typ, nullable, column_descr

    For a Django model, many of these are unknown. We'll fill them with blanks or placeholders:
      schema_nm -> "public" (placeholder)
      table_nm -> the model class name
      obj_typ -> "TBL"
      ord -> (blank)
      is_key -> (blank)
      column_nm -> field name
      data_typ -> (blank or "CharField"? Hard to parse reliably. We'll set blank.)
      nullable -> (blank)
      column_descr -> help_text
    """
    # If we want to merge into an existing file, let's attempt to read it first:
    if excel_path.lower().endswith(".csv"):
        try:
            df = pd.read_csv(excel_path)
        except FileNotFoundError:
            df = pd.DataFrame(
                columns=[
                    "schema_nm",
                    "table_nm",
                    "obj_typ",
                    "ord",
                    "is_key",
                    "column_nm",
                    "data_typ",
                    "nullable",
                    "column_descr",
                ]
            )
    else:
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            # Convert to DataFrame for easier merges
            data = []
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(row)
            df = pd.DataFrame(data, columns=headers)
        except FileNotFoundError:
            df = pd.DataFrame(
                columns=[
                    "schema_nm",
                    "table_nm",
                    "obj_typ",
                    "ord",
                    "is_key",
                    "column_nm",
                    "data_typ",
                    "nullable",
                    "column_descr",
                ]
            )

    # Build a list of new rows from help_dict
    new_rows = []
    for model_name, fields in help_dict.items():
        for field_name, help_txt in fields.items():
            new_rows.append(
                {
                    "schema_nm": "public",  # placeholder
                    "table_nm": model_name,
                    "obj_typ": "TBL",
                    "ord": "",
                    "is_key": "",
                    "column_nm": field_name,
                    "data_typ": "",
                    "nullable": "",
                    "column_descr": help_txt,
                }
            )

    # Convert new_rows to a DataFrame
    new_df = pd.DataFrame(new_rows, columns=df.columns)

    # Simple approach: just append new rows to existing data, possibly creating duplicates.
    # If you want to merge or update existing rows, you'll need more logic to match on (table_nm, column_nm).
    merged_df = pd.concat([df, new_df], ignore_index=True)

    # Write out
    if excel_path.lower().endswith(".csv"):
        merged_df.to_csv(excel_path, index=False)
        print(f"[INFO] Updated/created CSV file: {excel_path}")
    else:
        # Overwrite the Excel file
        wb = Workbook()
        ws = wb.active
        ws.append(list(merged_df.columns))  # header
        for row_data in merged_df.itertuples(index=False):
            ws.append(list(row_data))
        wb.save(excel_path)
        print(f"[INFO] Updated/created Excel file: {excel_path}")


################################################################################
# 7) AST Utility Functions
################################################################################


def is_django_model_class(class_node):
    """Check is django model."""
    for base in class_node.bases:
        if (isinstance(base, ast.Attribute) and base.attr == "Model") or (
            isinstance(base, ast.Name) and base.id == "Model"
        ):
            return True
    return False


def add_help_text_if_missing(call_node, new_help_text):
    """Add help text."""
    if not isinstance(call_node, ast.Call):
        return
    for kw in call_node.keywords:
        if kw.arg == "help_text":
            # Already set
            return
    call_node.keywords.append(
        ast.keyword(arg="help_text", value=ast.Constant(value=new_help_text))
    )


def get_help_text_from_field(call_node):
    """Get help text."""
    if isinstance(call_node, ast.Call):
        for kw in call_node.keywords:
            if kw.arg == "help_text":
                if isinstance(kw.value, ast.Constant):
                    return kw.value.value
    return None


################################################################################
# 8) CLI Entry Point
################################################################################

if __name__ == "__main__":
    os.chdir(sys.path[0] + "/..")
    print("CWD: {}".format(os.getcwd()))
    print("CWD: {}".format(sys.path[0]))
    parser = argparse.ArgumentParser(
        description="Synchronize Django model help_text with Excel/CSV or Postgres data dictionary."
    )
    parser.add_argument(
        "--models",
        required=True,
        help="Path to Django models.py (unused if --fromSql or --toSql only).",
    )
    parser.add_argument(
        "--excel",
        required=True,
        help="Path to the Excel or CSV file. In fromSql mode, this is CSV output. In other modes, read/write.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="If fromExcel, specify where to write updated models. Else we overwrite the original models file.",
    )

    group = parser.add_mutually_exclusive_group()
    group.add_argument(
        "--fromExcel",
        action="store_true",
        help="Update the Django source from Excel/CSV. (default)",
    )
    group.add_argument(
        "--toExcel",
        action="store_true",
        help="Update or create the Excel/CSV from the Django source.",
    )
    group.add_argument(
        "--fromSql",
        action="store_true",
        help="Run the built-in SQL script against Postgres and save to CSV.",
    )
    group.add_argument(
        "--toSql",
        action="store_true",
        help="Update Postgres data dictionary from Excel/CSV (COMMENT ON COLUMN).",
    )

    # Default behaviors
    parser.set_defaults(fromExcel=False, toExcel=False, fromSql=False, toSql=False)

    # Database connection parameters
    parser.add_argument("--db-host", default="localhost", help="PostgreSQL host.")
    parser.add_argument("--db-port", default="5432", help="PostgreSQL port.")
    parser.add_argument(
        "--db-name", default="postgres", help="PostgreSQL database name."
    )
    parser.add_argument("--db-user", default="postgres", help="PostgreSQL user.")
    parser.add_argument("--db-password", default="", help="PostgreSQL password.")
    parser.add_argument(
        "--db-schema",
        default="public",
        help="PostgreSQL schema name to use. Default is 'public'.",
    )

    args = parser.parse_args()
    output_path = args.output if args.output else args.models

    main(
        models_path=args.models,
        excel_path=args.excel,
        output_path=output_path,
        from_excel=args.fromExcel,
        to_excel=args.toExcel,
        from_sql=args.fromSql,
        to_sql=args.toSql,
        db_host=args.db_host,
        db_port=args.db_port,
        db_name=args.db_name,
        db_user=args.db_user,
        db_password=args.db_password,
        db_schema=args.db_schema,
    )
